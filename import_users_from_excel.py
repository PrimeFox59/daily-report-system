"""Import users from Excel file to database."""
import openpyxl
from app import create_app
from models import db, User


def import_users_from_excel(excel_file='User parin.xlsx', sheet_name='ID karyawan'):
    """
    Import users from Excel file to database.
    
    Expected Excel columns:
    A: id (employee_id)
    B: password
    C: fullname
    D: departemen
    E: seksi
    F: nama jabatan
    G: status
    """
    app = create_app()
    with app.app_context():
        # Load the Excel file
        try:
            wb = openpyxl.load_workbook(excel_file)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                # If sheet name not found, use the first sheet
                ws = wb.active
                print(f"Sheet '{sheet_name}' not found. Using '{ws.title}' instead.")
            
            # Clear existing users (optional - comment out if you want to keep existing users)
            db.drop_all()
            db.create_all()
            print("Database cleared and recreated.")
            
            # Skip header row (row 1)
            imported_count = 0
            skipped_count = 0
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Extract data from columns
                employee_id = str(row[0]).strip() if row[0] else None
                password = str(row[1]).strip() if row[1] else 'zzz'
                fullname = str(row[2]).strip() if row[2] else None
                departemen = str(row[3]).strip() if row[3] else ''
                seksi = str(row[4]).strip() if row[4] else ''
                nama_jabatan = str(row[5]).strip() if row[5] else 'user'
                status = str(row[6]).strip() if row[6] else 'approved'
                
                # Skip if employee_id or fullname is missing
                if not employee_id or not fullname or employee_id == 'None':
                    skipped_count += 1
                    continue
                
                # Check if status is approved
                if status.lower() != 'approved':
                    print(f"Skipping {fullname} (ID: {employee_id}) - Status: {status}")
                    skipped_count += 1
                    continue
                
                # Check if user already exists
                existing_user = User.query.filter_by(employee_id=employee_id).first()
                if existing_user:
                    print(f"User {employee_id} already exists, skipping...")
                    skipped_count += 1
                    continue
                
                # Create new user
                # First user (ID 344) will be admin
                is_admin = (employee_id == '344')
                
                user = User(
                    employee_id=employee_id,
                    name=fullname,
                    department=departemen,
                    section=seksi,
                    job=nama_jabatan,
                    shift='Pagi',  # Default shift, can be updated later
                    is_admin=is_admin
                )
                user.set_password(password)
                
                db.session.add(user)
                imported_count += 1
                
                if is_admin:
                    print(f"✓ Imported ADMIN: {fullname} (ID: {employee_id})")
                else:
                    print(f"✓ Imported: {fullname} (ID: {employee_id})")
            
            # Commit all changes
            db.session.commit()
            
            print(f"\n{'='*60}")
            print(f"Import completed!")
            print(f"Total imported: {imported_count} users")
            print(f"Total skipped: {skipped_count} users")
            print(f"{'='*60}")
            
            # Show summary
            total_users = User.query.count()
            admin_users = User.query.filter_by(is_admin=True).count()
            print(f"\nDatabase now contains:")
            print(f"  - Total users: {total_users}")
            print(f"  - Admin users: {admin_users}")
            print(f"  - Regular users: {total_users - admin_users}")
            
        except FileNotFoundError:
            print(f"Error: File '{excel_file}' not found!")
            print("Make sure 'User parin.xlsx' is in the same directory as this script.")
        except Exception as e:
            print(f"Error during import: {str(e)}")
            db.session.rollback()


if __name__ == '__main__':
    import_users_from_excel()
